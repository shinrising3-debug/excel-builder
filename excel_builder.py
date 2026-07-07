"""
╔══════════════════════════════════════════════════════════╗
║          EXCEL BUILDER  v1.0  —  by TTR                  ║
║   Describe any spreadsheet in plain English,             ║
║   and this tool builds it for you.                       ║
╚══════════════════════════════════════════════════════════╝

Modes:
  [1] AI Mode   — describe what you want, Claude builds it (needs API key)
  [2] Manual    — step-by-step column builder (no API needed)

Setup:
  - Put your Anthropic API key in a file called '.api_key' in the same
    folder, OR set the environment variable ANTHROPIC_API_KEY.
  - First run auto-installs openpyxl and anthropic if missing.

Run via: double-click Excel_Builder.bat
"""

import os, sys, json, re, textwrap, traceback
from datetime import datetime

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
OUTPUT_DIR = SCRIPT_DIR

# ── Auto-install deps ──
def ensure_packages():
    required = ["openpyxl", "anthropic"]
    missing = []
    for pkg in required:
        try:
            __import__(pkg)
        except ImportError:
            missing.append(pkg)
    if missing:
        print(f"\n  [!] Installing: {', '.join(missing)}...\n")
        os.system(f'"{sys.executable}" -m pip install {" ".join(missing)} --quiet')
        print("\n  [✓] Done. Restarting...\n")
        os.execv(sys.executable, [sys.executable] + sys.argv)

ensure_packages()

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter


# ═══════════════════════════════════════════════════════
#  HELPERS
# ═══════════════════════════════════════════════════════
def clear():
    os.system('cls' if os.name == 'nt' else 'clear')

def ask(prompt, default=None):
    suffix = f" [{default}]" if default else ""
    val = input(f"  {prompt}{suffix}: ").strip()
    return val if val else default

def get_api_key():
    key = os.environ.get("ANTHROPIC_API_KEY", "").strip()
    if key:
        return key
    key_file = os.path.join(SCRIPT_DIR, ".api_key")
    if os.path.exists(key_file):
        with open(key_file, "r") as f:
            key = f.read().strip()
        if key:
            return key
    return None

def sanitize_filename(name):
    name = re.sub(r'[<>:"/\\|?*]', '', name)
    name = name.strip().replace(' ', '_')
    if not name.endswith('.xlsx'):
        name += '.xlsx'
    return name


# ═══════════════════════════════════════════════════════
#  AI MODE — Claude generates the spreadsheet
# ═══════════════════════════════════════════════════════

SYSTEM_PROMPT = """You are an Excel spreadsheet generator. The user will describe a spreadsheet they need.

You must respond with ONLY valid Python code that uses openpyxl to create the spreadsheet.
No explanation, no markdown, no ```python blocks — JUST the raw Python code.

RULES:
1. Always save to the path stored in variable OUTPUT_PATH (it will be defined before your code runs)
2. Always use the variable FILE_NAME for the filename (also pre-defined)
3. Import everything you need at the top
4. Make it look PROFESSIONAL and PRESENTABLE:
   - Turn off gridlines: ws.sheet_view.showGridLines = False
   - Use 'Aptos' font throughout
   - Dark header row (background: 1B2A4A, font: white, bold)
   - Gold accent line under headers (bottom border: D4AF37, medium)
   - Cell borders on ALL data cells (thin, color CED4DA)
   - Alternating row colors (white FFFFFF and off-white F8F9FA)
   - Proper column widths (auto-size based on content type)
   - Freeze the header row
   - All body text must be BLACK (color: 212529)
   - Wrap text and center-align short fields, left-align long text
   - Row heights: headers 30, data rows 28
   - Add a title row at top (merged, font size 16, bold)
   - Add a subtle subtitle row below title
   - Add a 4px gold accent bar as row 1
   - Data starts after the header row
5. Add data validation dropdowns where appropriate (showDropDown=False to show the arrow)
6. Add conditional formatting with BACKGROUND color only (no font color change) where it makes sense
7. Pre-format at least 200 empty rows with borders, fonts, alignment, and alternating fill
8. Use PatternFill with start_color and end_color with FF prefix for conditional formatting
9. Set landscape orientation and fit to page width
10. Include any sample data the user mentioned
11. If the user mentions specific columns, use EXACTLY those column names
12. For date columns use number_format = "mm/dd/yyyy"
13. For currency columns use number_format = '$#,##0.00'
14. NEVER use Excel Tables (ws.add_table) — they cause rendering issues

The output file path will be: OUTPUT_PATH (pre-defined variable, do not redefine it)

Your code will be exec'd in a Python environment with openpyxl already imported.
OUTPUT_PATH and FILE_NAME will already be defined as variables before your code runs.
"""

def ai_mode():
    clear()
    api_key = get_api_key()

    if not api_key:
        print("\n  ╔════════════════════════════════════════════════╗")
        print("  ║  API key not found. To use AI mode:            ║")
        print("  ║                                                ║")
        print("  ║  Option A: Create a file called '.api_key'     ║")
        print("  ║  in the same folder as this script and paste   ║")
        print("  ║  your Anthropic API key inside.                ║")
        print("  ║                                                ║")
        print("  ║  Option B: Set environment variable            ║")
        print("  ║  ANTHROPIC_API_KEY=sk-ant-...                  ║")
        print("  ╚════════════════════════════════════════════════╝\n")
        input("  Press Enter to go back...")
        return

    import anthropic
    client = anthropic.Anthropic(api_key=api_key)

    print("\n  ── AI EXCEL BUILDER ────────────────────────────")
    print("  Describe the spreadsheet you need in plain English.")
    print("  Be as detailed as you want — columns, dropdowns,")
    print("  sample data, formatting, anything.\n")
    print("  Examples:")
    print("    • 'Employee attendance tracker with name, date,")
    print("       clock in, clock out, hours, and overtime columns'")
    print("    • 'Inventory sheet for warehouse with SKU, item,")
    print("       quantity, location, reorder point, status dropdown'")
    print("    • 'Monthly budget tracker with income and expense")
    print("       categories, amounts, and running balance'\n")
    print("  ────────────────────────────────────────────────\n")

    # Multi-line input
    print("  Describe your spreadsheet (press Enter twice when done):\n")
    lines = []
    empty_count = 0
    while True:
        try:
            line = input("  > ")
            if line.strip() == "":
                empty_count += 1
                if empty_count >= 1 and lines:
                    break
            else:
                empty_count = 0
                lines.append(line)
        except EOFError:
            break

    if not lines:
        print("  No description provided.")
        return

    description = "\n".join(lines)

    # Get filename
    file_name = ask("\n  File name", "My_Spreadsheet")
    file_name = sanitize_filename(file_name)
    output_path = os.path.join(OUTPUT_DIR, file_name)

    print(f"\n  ⏳ Generating '{file_name}'... (this takes 10-30 seconds)\n")

    try:
        response = client.messages.create(
            model="claude-sonnet-4-6",
            max_tokens=8000,
            system=SYSTEM_PROMPT,
            messages=[
                {"role": "user", "content": description}
            ]
        )

        code = response.content[0].text

        # Clean up — strip markdown fences if Claude added them despite instructions
        code = re.sub(r'^```python\s*\n?', '', code, flags=re.MULTILINE)
        code = re.sub(r'^```\s*$', '', code, flags=re.MULTILINE)
        code = code.strip()

        # Save generated code for debugging
        code_log = os.path.join(SCRIPT_DIR, "_last_generated_code.py")
        with open(code_log, "w", encoding="utf-8") as f:
            f.write(f"# Generated: {datetime.now()}\n")
            f.write(f"# Description: {description[:200]}\n\n")
            f.write(code)

        # Execute the generated code
        exec_globals = {
            "OUTPUT_PATH": output_path,
            "FILE_NAME": file_name,
            "__builtins__": __builtins__,
        }

        exec(code, exec_globals)

        if os.path.exists(output_path):
            size = os.path.getsize(output_path)
            print(f"  ╔════════════════════════════════════════════════╗")
            print(f"  ║  ✓  FILE CREATED SUCCESSFULLY                 ║")
            print(f"  ╠════════════════════════════════════════════════╣")
            print(f"  ║  File: {file_name:<40} ║")
            print(f"  ║  Size: {size:,} bytes{' '*(34-len(f'{size:,} bytes'))} ║")
            print(f"  ║  Path: {OUTPUT_DIR[:40]:<40} ║")
            print(f"  ╚════════════════════════════════════════════════╝")

            open_it = ask("\n  Open in Excel now? (Y/n)", "Y")
            if open_it.upper() == "Y":
                if os.name == 'nt':
                    os.startfile(output_path)
                elif sys.platform == 'darwin':
                    os.system(f'open "{output_path}"')
                else:
                    os.system(f'xdg-open "{output_path}"')
        else:
            print("  [!] Code ran but file was not created.")
            print(f"  Check {code_log} for the generated code.")

    except Exception as e:
        print(f"\n  [!] Error: {e}")
        print(f"\n  Generated code saved to: {code_log}")
        traceback.print_exc()

    input("\n  Press Enter to continue...")


# ═══════════════════════════════════════════════════════
#  MANUAL MODE — step-by-step builder
# ═══════════════════════════════════════════════════════

def manual_mode():
    clear()
    print("\n  ── MANUAL EXCEL BUILDER ────────────────────────")
    print("  Build a formatted spreadsheet step by step.\n")

    title = ask("Sheet title", "My Tracker")
    file_name = sanitize_filename(ask("File name", title))
    output_path = os.path.join(OUTPUT_DIR, file_name)

    # Collect columns
    print("\n  Add your columns one at a time.")
    print("  For each column, pick a type that controls formatting.")
    print("  Type 'done' when finished.\n")

    columns = []
    col_types = ["Text", "Number", "Currency", "Date", "Dropdown", "Long Text"]

    while True:
        name = ask(f"  Column {len(columns)+1} name (or 'done')")
        if not name or name.lower() == 'done':
            if len(columns) < 1:
                print("  Need at least 1 column.")
                continue
            break

        print(f"\n  Column type for '{name}':")
        for i, t in enumerate(col_types, 1):
            print(f"    [{i}] {t}")
        try:
            ct = int(ask("  Type #", "1"))
            col_type = col_types[ct - 1] if 1 <= ct <= len(col_types) else "Text"
        except (ValueError, TypeError):
            col_type = "Text"

        dropdown_items = None
        if col_type == "Dropdown":
            items = ask("  Dropdown options (comma-separated)")
            if items:
                dropdown_items = [x.strip() for x in items.split(",")]

        width = {"Text": 20, "Number": 12, "Currency": 14, "Date": 16,
                 "Dropdown": 18, "Long Text": 40}.get(col_type, 20)

        columns.append({
            "name": name,
            "type": col_type,
            "width": width,
            "dropdown": dropdown_items,
        })
        print(f"  ✓ Added: {name} ({col_type})\n")

    # How many pre-formatted rows?
    try:
        num_rows = int(ask("\n  How many blank rows to pre-format?", "200"))
    except (ValueError, TypeError):
        num_rows = 200
    num_rows = max(10, min(num_rows, 2000))

    print(f"\n  ⏳ Building '{file_name}'...\n")

    # ── Build the workbook ──
    DARK_BG   = "1B2A4A"
    GOLD      = "D4AF37"
    WHITE     = "FFFFFF"
    OFF_WHITE = "F8F9FA"
    MID_GRAY  = "CED4DA"
    TXT       = "212529"
    MUTED     = "6C757D"

    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]
    ws.sheet_view.showGridLines = False

    total_cols = len(columns)
    last_col = get_column_letter(total_cols)

    cell_border = Border(
        left=Side(style="thin", color=MID_GRAY),
        right=Side(style="thin", color=MID_GRAY),
        top=Side(style="thin", color=MID_GRAY),
        bottom=Side(style="thin", color=MID_GRAY),
    )

    for i, col in enumerate(columns, 1):
        ws.column_dimensions[get_column_letter(i)].width = col["width"]

    # Gold bar
    for c in range(1, total_cols + 1):
        ws.cell(row=1, column=c).fill = PatternFill("solid", fgColor=GOLD)
    ws.row_dimensions[1].height = 4

    # Title
    ws.merge_cells(f"A2:{last_col}2")
    ws["A2"].value = title.upper()
    ws["A2"].font = Font(name="Aptos", bold=True, size=16, color=DARK_BG)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[2].height = 38

    # Subtitle
    ws.merge_cells(f"A3:{last_col}3")
    ws["A3"].value = f"Generated {datetime.now().strftime('%b %d, %Y')}  ·  {num_rows} rows pre-formatted"
    ws["A3"].font = Font(name="Aptos", size=10, color=MUTED)
    ws["A3"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[3].height = 22
    ws.row_dimensions[4].height = 6

    # Headers
    header_row = 5
    for i, col in enumerate(columns, 1):
        c = ws.cell(row=header_row, column=i, value=col["name"])
        c.font = Font(name="Aptos", bold=True, size=10, color=WHITE)
        c.fill = PatternFill("solid", fgColor=DARK_BG)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = Border(
            left=Side(style="thin", color="0F1B33"),
            right=Side(style="thin", color="0F1B33"),
            top=Side(style="thin", color="0F1B33"),
            bottom=Side(style="medium", color=GOLD),
        )
    ws.row_dimensions[header_row].height = 30

    # Data rows
    data_start = 6
    center_types = {"Number", "Currency", "Date", "Dropdown"}

    for row in range(data_start, data_start + num_rows):
        is_alt = (row - data_start) % 2 == 1
        fill = PatternFill("solid", fgColor=OFF_WHITE) if is_alt else PatternFill("solid", fgColor=WHITE)
        for i, col in enumerate(columns, 1):
            c = ws.cell(row=row, column=i)
            c.font = Font(name="Aptos", size=10, color=TXT)
            c.border = cell_border
            h = "center" if col["type"] in center_types else "left"
            c.alignment = Alignment(horizontal=h, vertical="center", wrap_text=True)
            c.fill = fill
            if col["type"] == "Date":
                c.number_format = "mm/dd/yyyy"
            elif col["type"] == "Currency":
                c.number_format = '$#,##0.00'
        ws.row_dimensions[row].height = 28

    # Dropdowns
    end_row = data_start + num_rows - 1
    for i, col in enumerate(columns, 1):
        if col["type"] == "Dropdown" and col["dropdown"]:
            formula = '"' + ",".join(col["dropdown"]) + '"'
            dv = DataValidation(type="list", formula1=formula,
                                allow_blank=True, showDropDown=False)
            dv.showInputMessage = True
            letter = get_column_letter(i)
            dv.add(f"{letter}{data_start}:{letter}{end_row}")
            ws.add_data_validation(dv)

    ws.freeze_panes = f"A{data_start}"
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

    wb.save(output_path)

    print(f"  ╔════════════════════════════════════════════════╗")
    print(f"  ║  ✓  FILE CREATED SUCCESSFULLY                 ║")
    print(f"  ╠════════════════════════════════════════════════╣")
    print(f"  ║  File: {file_name:<40} ║")
    print(f"  ║  Columns: {total_cols:<37} ║")
    print(f"  ║  Rows: {num_rows:<39} ║")
    print(f"  ╚════════════════════════════════════════════════╝")

    open_it = ask("\n  Open in Excel now? (Y/n)", "Y")
    if open_it.upper() == "Y":
        if os.name == 'nt':
            os.startfile(output_path)
        elif sys.platform == 'darwin':
            os.system(f'open "{output_path}"')
        else:
            os.system(f'xdg-open "{output_path}"')

    input("\n  Press Enter to continue...")


# ═══════════════════════════════════════════════════════
#  MAIN
# ═══════════════════════════════════════════════════════
def main():
    while True:
        clear()
        has_key = get_api_key() is not None
        key_status = "✓ API key found" if has_key else "✗ No API key"

        print()
        print("  ╔══════════════════════════════════════════════════╗")
        print("  ║          EXCEL BUILDER  v1.0                    ║")
        print("  ║          Build any spreadsheet, anytime.        ║")
        print("  ╠══════════════════════════════════════════════════╣")
        print(f"  ║  [{key_status}]{' '*(35-len(key_status))}  ║")
        print("  ║                                                 ║")
        print("  ║  [1]  AI Mode — describe it, Claude builds it   ║")
        print("  ║  [2]  Manual  — step-by-step column builder     ║")
        print("  ║  [3]  Set API key                               ║")
        print("  ║  [4]  Exit                                      ║")
        print("  ╚══════════════════════════════════════════════════╝")
        print()

        choice = ask("Pick an option", "1")

        if choice == "1":
            ai_mode()
        elif choice == "2":
            manual_mode()
        elif choice == "3":
            clear()
            print("\n  ── SET API KEY ─────────────────────────────────")
            print("  Get your key from: console.anthropic.com\n")
            key = ask("Paste your Anthropic API key (sk-ant-...)")
            if key and key.startswith("sk-"):
                key_file = os.path.join(SCRIPT_DIR, ".api_key")
                with open(key_file, "w") as f:
                    f.write(key)
                print(f"\n  ✓ Key saved to .api_key")
            else:
                print("  Invalid key format.")
            input("\n  Press Enter to continue...")
        elif choice == "4":
            print("\n  Later. ✌️\n")
            break
        else:
            pass


if __name__ == "__main__":
    main()
