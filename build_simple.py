"""
Simple Form Excel Builder — reads guided inputs from GitHub Actions form.
No AI, no API, completely free.
"""

import os, sys, re
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter

DARK_BG   = "1B2A4A"
GOLD      = "D4AF37"
WHITE     = "FFFFFF"
OFF_WHITE = "F8F9FA"
MID_GRAY  = "CED4DA"
TXT       = "212529"
MUTED     = "6C757D"

CF_COLORS = ["FFD1E7DD","FFFFF3CD","FFF8D7DA","FFCFE2FF","FFE2D9F3","FFFDE8CD","FFD4EDDA","FFFFF0F0"]

def border():
    return Border(
        left=Side(style="thin", color=MID_GRAY), right=Side(style="thin", color=MID_GRAY),
        top=Side(style="thin", color=MID_GRAY), bottom=Side(style="thin", color=MID_GRAY),
    )

def main():
    title      = os.environ.get("INPUT_TITLE", "My Sheet").strip()
    status     = os.environ.get("INPUT_STATUS", "No").strip()
    status_opts= os.environ.get("INPUT_STATUS_OPTS", "").strip()
    cols_raw   = os.environ.get("INPUT_COLUMNS", "").strip()
    types_raw  = os.environ.get("INPUT_TYPES", "").strip()
    filename   = os.environ.get("INPUT_FILENAME", "My_Spreadsheet").strip()

    try:
        num_rows = int(os.environ.get("INPUT_ROWS", "200"))
    except ValueError:
        num_rows = 200
    num_rows = max(10, min(num_rows, 2000))

    if not filename.endswith('.xlsx'):
        filename = re.sub(r'[<>:"/\\|?*]', '', filename).replace(' ', '_') + '.xlsx'

    if not cols_raw:
        print("❌ No columns provided.")
        sys.exit(1)

    # Parse columns
    col_names = [c.strip() for c in cols_raw.split(',') if c.strip()]
    col_types_list = [t.strip().lower() for t in types_raw.split(',')] if types_raw else []

    # Pad types to match columns
    while len(col_types_list) < len(col_names):
        col_types_list.append('text')

    # Build column defs
    columns = []
    for name, ctype in zip(col_names, col_types_list):
        if ctype not in ('text', 'date', 'number', 'currency', 'long'):
            ctype = 'text'
        width = {'text': 22, 'date': 16, 'number': 14, 'currency': 16, 'long': 42}.get(ctype, 22)
        columns.append({'name': name, 'type': ctype, 'width': width})

    # Add status column if requested
    add_status = status.lower() == 'yes' and status_opts
    status_options = []
    if add_status:
        status_options = [o.strip() for o in status_opts.split('|') if o.strip()]
        columns.append({'name': 'Status', 'type': 'dropdown', 'width': 20})

    # Always prepend #
    all_columns = [{'name': '#', 'type': 'number', 'width': 5}] + columns
    total_cols = len(all_columns)
    last_col = get_column_letter(total_cols)
    center_types = {'number', 'currency', 'date', 'dropdown'}

    print(f"📊 Building: {filename}")
    print(f"📋 Title: {title}")
    print(f"📐 Columns ({len(columns)}):")
    for col in columns:
        extra = f" → {', '.join(status_options)}" if col['type'] == 'dropdown' else ""
        print(f"   • {col['name']} ({col['type']}){extra}")
    if add_status:
        print(f"🎨 Status filter: {', '.join(status_options)} (full row coloring)")
    print(f"📏 Rows: {num_rows}")
    print()

    # ── Build ──
    wb = Workbook()
    ws = wb.active
    # Sanitize sheet title (Excel doesn't allow : \ / ? * [ ] in sheet names)
    safe_title = re.sub(r'[:\\/?*\[\]]', '-', title)
    ws.title = safe_title[:31]
    ws.sheet_view.showGridLines = False

    for i, col in enumerate(all_columns, 1):
        ws.column_dimensions[get_column_letter(i)].width = col['width']

    # Gold bar
    for c in range(1, total_cols + 1):
        ws.cell(row=1, column=c).fill = PatternFill("solid", fgColor=GOLD)
    ws.row_dimensions[1].height = 4

    # Title
    ws.merge_cells(f"A2:{last_col}2")
    ws["A2"].value = title.upper()
    ws["A2"].font = Font(name="Aptos", bold=True, size=16, color=DARK_BG)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[2].height = 38

    # Subtitle
    ws.merge_cells(f"A3:{last_col}3")
    sub = f"Generated {datetime.now().strftime('%b %d, %Y')}  ·  {num_rows} rows pre-formatted"
    if add_status:
        sub += f"  ·  Status dropdown colors entire row"
    ws["A3"].value = sub
    ws["A3"].font = Font(name="Aptos", size=10, color=MUTED)
    ws["A3"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[3].height = 22
    ws.row_dimensions[4].height = 6

    # Headers
    header_row = 5
    for i, col in enumerate(all_columns, 1):
        c = ws.cell(row=header_row, column=i, value=col['name'])
        c.font = Font(name="Aptos", bold=True, size=10, color=WHITE)
        c.fill = PatternFill("solid", fgColor=DARK_BG)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = Border(
            left=Side(style="thin", color="0F1B33"), right=Side(style="thin", color="0F1B33"),
            top=Side(style="thin", color="0F1B33"), bottom=Side(style="medium", color=GOLD),
        )
    ws.row_dimensions[header_row].height = 30

    # Data rows
    data_start = 6
    end_row = data_start + num_rows - 1

    for row in range(data_start, end_row + 1):
        is_alt = (row - data_start) % 2 == 1
        fill = PatternFill("solid", fgColor=OFF_WHITE) if is_alt else PatternFill("solid", fgColor=WHITE)
        for i, col in enumerate(all_columns, 1):
            c = ws.cell(row=row, column=i)
            c.font = Font(name="Aptos", size=10, color=TXT)
            c.border = border()
            h = "center" if col['type'] in center_types or col['name'] == '#' else "left"
            c.alignment = Alignment(horizontal=h, vertical="center", wrap_text=True)
            c.fill = fill
            if col['type'] == 'date':
                c.number_format = "mm/dd/yyyy"
            elif col['type'] == 'currency':
                c.number_format = '$#,##0.00'
        ws.row_dimensions[row].height = 28

    # Status dropdown + conditional formatting
    if add_status and status_options:
        status_col_idx = total_cols  # last column
        status_letter = get_column_letter(status_col_idx)
        formula = '"' + ",".join(status_options) + '"'
        dv = DataValidation(type="list", formula1=formula, allow_blank=True, showDropDown=False)
        dv.prompt = "Select status"
        dv.showInputMessage = True
        dv.showErrorMessage = True
        dv.add(f"{status_letter}{data_start}:{status_letter}{end_row}")
        ws.add_data_validation(dv)

        # Full row conditional formatting
        full_range = f"A{data_start}:{last_col}{end_row}"
        for idx, option in enumerate(status_options):
            color = CF_COLORS[idx % len(CF_COLORS)]
            ws.conditional_formatting.add(full_range, FormulaRule(
                formula=[f'${status_letter}{data_start}="{option}"'],
                fill=PatternFill(start_color=color, end_color=color, fill_type="solid"),
                stopIfTrue=True,
            ))

    # Freeze + print
    ws.freeze_panes = f"A{data_start}"
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

    output = os.path.join(os.getcwd(), filename)
    wb.save(output)
    size = os.path.getsize(output)
    print(f"✅ Done — {filename} ({size:,} bytes)")


if __name__ == "__main__":
    main()
