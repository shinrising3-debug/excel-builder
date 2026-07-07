"""
Pure Python Excel Generator — No AI, No API keys, completely free.
Parses column definitions from workflow input and builds a formatted .xlsx.
"""

import os, sys, re
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter

# ── Palette ──
DARK_BG   = "1B2A4A"
GOLD      = "D4AF37"
WHITE     = "FFFFFF"
OFF_WHITE = "F8F9FA"
MID_GRAY  = "CED4DA"
TXT       = "212529"
MUTED     = "6C757D"

CF_COLORS = [
    "FFD1E7DD",  # green
    "FFFFF3CD",  # amber
    "FFF8D7DA",  # red
    "FFCFE2FF",  # blue
    "FFE2D9F3",  # purple
    "FFFDE8CD",  # orange
    "FFD4EDDA",  # mint
    "FFFFF0F0",  # pink
]


def get_border():
    return Border(
        left=Side(style="thin", color=MID_GRAY),
        right=Side(style="thin", color=MID_GRAY),
        top=Side(style="thin", color=MID_GRAY),
        bottom=Side(style="thin", color=MID_GRAY),
    )


def parse_columns(raw):
    """
    Parse column definitions. Format:
      Name, Date (date), Amount (currency), Status (dropdown: Pending | Fixed | Escalated), Notes (long)

    Types: text (default), date, number, currency, dropdown, long
    Dropdown options after colon, separated by |
    """
    columns = []
    parts = re.split(r',\s*(?![^(]*\))', raw)

    for part in parts:
        part = part.strip()
        if not part:
            continue

        match = re.match(r'^(.+?)\s*\(\s*(.*?)\s*\)\s*$', part)

        if match:
            name = match.group(1).strip()
            type_str = match.group(2).strip().lower()

            if type_str.startswith('dropdown'):
                col_type = 'dropdown'
                opts_match = re.search(r':\s*(.+)', type_str)
                options = [o.strip() for o in opts_match.group(1).split('|')] if opts_match else ["Option 1", "Option 2", "Option 3"]
            else:
                col_type = type_str if type_str in ('date', 'number', 'currency', 'long', 'text') else 'text'
                options = None
        else:
            name = part.strip()
            col_type = 'text'
            options = None

        width_map = {'text': 22, 'number': 14, 'currency': 16, 'date': 16, 'dropdown': 20, 'long': 42}
        columns.append({'name': name, 'type': col_type, 'width': width_map.get(col_type, 22), 'options': options})

    return columns


def build_excel(title, columns, num_rows, filename, sample_data=None):
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]
    ws.sheet_view.showGridLines = False

    all_columns = [{'name': '#', 'type': 'number', 'width': 5, 'options': None}] + columns
    total_cols = len(all_columns)
    last_col = get_column_letter(total_cols)
    center_types = {'number', 'currency', 'date', 'dropdown'}

    for i, col in enumerate(all_columns, 1):
        ws.column_dimensions[get_column_letter(i)].width = col['width']

    # Gold bar
    for c in range(1, total_cols + 1):
        ws.cell(row=1, column=c).fill = PatternFill("solid", fgColor=GOLD)
    ws.row_dimensions[1].height = 4

    # Title
    ws.merge_cells(f"A2:{last_col}2")
    ws["A2"].value = title.upper()
    ws["A2"].font = Font(name="Aptos", bold=True, size=16, color=DARK_BG)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[2].height = 38

    # Subtitle
    ws.merge_cells(f"A3:{last_col}3")
    ws["A3"].value = f"Generated {datetime.now().strftime('%b %d, %Y')}  ·  {num_rows} rows pre-formatted"
    ws["A3"].font = Font(name="Aptos", size=10, color=MUTED)
    ws["A3"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[3].height = 22
    ws.row_dimensions[4].height = 6

    # Headers
    header_row = 5
    for i, col in enumerate(all_columns, 1):
        c = ws.cell(row=header_row, column=i, value=col['name'])
        c.font = Font(name="Aptos", bold=True, size=10, color=WHITE)
        c.fill = PatternFill("solid", fgColor=DARK_BG)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = Border(
            left=Side(style="thin", color="0F1B33"), right=Side(style="thin", color="0F1B33"),
            top=Side(style="thin", color="0F1B33"), bottom=Side(style="medium", color=GOLD),
        )
    ws.row_dimensions[header_row].height = 30

    # Data rows
    data_start = 6
    end_row = data_start + num_rows - 1

    for row in range(data_start, end_row + 1):
        is_alt = (row - data_start) % 2 == 1
        fill = PatternFill("solid", fgColor=OFF_WHITE) if is_alt else PatternFill("solid", fgColor=WHITE)
        for i, col in enumerate(all_columns, 1):
            c = ws.cell(row=row, column=i)
            c.font = Font(name="Aptos", size=10, color=TXT)
            c.border = get_border()
            h = "center" if col['type'] in center_types or col['name'] == '#' else "left"
            c.alignment = Alignment(horizontal=h, vertical="center", wrap_text=True)
            c.fill = fill
            if col['type'] == 'date':
                c.number_format = "mm/dd/yyyy"
            elif col['type'] == 'currency':
                c.number_format = '$#,##0.00'
        ws.row_dimensions[row].height = 28

    # Sample data
    if sample_data:
        lines = [l.strip() for l in sample_data.strip().split('\n') if l.strip()]
        for row_offset, line in enumerate(lines):
            row = data_start + row_offset
            if row > end_row:
                break
            values = [v.strip() for v in line.split('|')]
            ws.cell(row=row, column=1, value=row_offset + 1)
            for col_idx, val in enumerate(values):
                if col_idx + 2 > total_cols:
                    break
                ws.cell(row=row, column=col_idx + 2, value=val)
            ws.row_dimensions[row].height = 34

    # Dropdowns
    for i, col in enumerate(all_columns, 1):
        if col['type'] == 'dropdown' and col.get('options'):
            formula = '"' + ",".join(col['options']) + '"'
            dv = DataValidation(type="list", formula1=formula, allow_blank=True, showDropDown=False)
            dv.prompt = f"Select {col['name'].lower()}"
            dv.showInputMessage = True
            dv.showErrorMessage = True
            letter = get_column_letter(i)
            dv.add(f"{letter}{data_start}:{letter}{end_row}")
            ws.add_data_validation(dv)

    # Conditional formatting on last dropdown column
    dropdown_cols = [(i, col) for i, col in enumerate(all_columns, 1) if col['type'] == 'dropdown' and col.get('options')]
    if dropdown_cols:
        cf_col_idx, cf_col = dropdown_cols[-1]
        cf_letter = get_column_letter(cf_col_idx)
        full_range = f"A{data_start}:{last_col}{end_row}"
        for opt_idx, option in enumerate(cf_col['options']):
            color = CF_COLORS[opt_idx % len(CF_COLORS)]
            ws.conditional_formatting.add(full_range, FormulaRule(
                formula=[f'${cf_letter}{data_start}="{option}"'],
                fill=PatternFill(start_color=color, end_color=color, fill_type="solid"),
                stopIfTrue=True,
            ))

    # Freeze + print
    ws.freeze_panes = f"A{data_start}"
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

    wb.save(filename)
    return filename


def main():
    title = os.environ.get("TITLE", "My Tracker").strip()
    columns_raw = os.environ.get("COLUMNS", "").strip()
    filename = os.environ.get("FILE_NAME", "My_Spreadsheet").strip()
    sample_data = os.environ.get("SAMPLE_DATA", "").strip()

    try:
        num_rows = int(os.environ.get("NUM_ROWS", "200"))
    except ValueError:
        num_rows = 200
    num_rows = max(10, min(num_rows, 2000))

    if not filename.endswith('.xlsx'):
        filename = re.sub(r'[<>:"/\\|?*]', '', filename).replace(' ', '_') + '.xlsx'

    if not columns_raw:
        print("❌ No columns provided.")
        print('   Example: Name, Date (date), Amount (currency), Status (dropdown: Pending|Fixed|Done)')
        sys.exit(1)

    columns = parse_columns(columns_raw)
    if not columns:
        print("❌ Could not parse columns.")
        sys.exit(1)

    print(f"📊 Building: {filename}")
    print(f"📋 Title: {title}")
    print(f"📐 Columns ({len(columns)}):")
    for col in columns:
        opts = f" → {', '.join(col['options'])}" if col.get('options') else ""
        print(f"   • {col['name']} ({col['type']}){opts}")
    print(f"📏 Rows: {num_rows}")
    print()

    output_path = os.path.join(os.getcwd(), filename)
    build_excel(title, columns, num_rows, output_path, sample_data or None)

    size = os.path.getsize(output_path)
    print(f"✅ Done — {filename} ({size:,} bytes)")


if __name__ == "__main__":
    main()
